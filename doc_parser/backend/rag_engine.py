#rag_engine.py
from .config import DATA_DIR, OPENAI_API_KEY, EMBEDDING_MODEL, LLM_MODEL, BATCH_SIZE
from .chroma import ChromaManager
from openai import OpenAI
from pathlib import Path
import re
import openpyxl
import pandas as pd
import openpyxl
from .utils.parser import parse_competition_query
from .utils.search_excel import find_competition_ratio
from .utils.generate_graph import generate_base64_graph

# OpenAI 클라이언트 초기화
openai_client = OpenAI(api_key=OPENAI_API_KEY)

# Chroma 매니저 초기화
chroma = ChromaManager()

# 문단 쪼개기 관련 상수
CHUNK_MAX_CHARS = 800
CHUNK_OVERLAP = 100


def split_into_chunks(text: str):
    """문장을 일정 길이로 나눔"""
    text = text.strip()
    if not text:
        return []
    if len(text) <= CHUNK_MAX_CHARS:
        return [text]

    chunks = []
    start = 0
    while start < len(text):
        end = start + CHUNK_MAX_CHARS
        chunk = text[start:end].strip()
        chunks.append(chunk)
        if end >= len(text):
            break
        start = end - CHUNK_OVERLAP
    return chunks


def _normalize_for_search(s: str) -> str:
    """검색용으로 텍스트 정규화: None 안전 처리, 소문자화, 모든 공백 제거."""
    if s is None:
        return ""
    # strip and collapse unicode whitespace, then remove all whitespace for robust matching
    s = re.sub(r"\s+", "", str(s))
    return s.lower()


def index_all_documents():
    """output 폴더의 모든 .md 및 .xlsx 문서를 읽고 Chroma에 임베딩 (중복 문서는 건너뜀)"""
    md_files = list(DATA_DIR.glob("*.md"))
    xlsx_files = list(DATA_DIR.glob("*.xlsx"))
    ids, texts, metadatas = [], [], []
    idx = 0

    # 이미 인덱싱된 문서 이름 목록 불러오기
    try:
        existing = chroma.collection.get(include=['metadatas'])
        existing_docs = set(meta['source'] for meta in existing['metadatas'] if meta and 'source' in meta)
    except Exception as e:
        print(f"⚠️ 기존 인덱싱 목록 불러오기 실패: {e}")
        existing_docs = set()

    # 새 문서만 인덱싱
    all_files = md_files + xlsx_files
    new_docs = [f for f in all_files if f.name not in existing_docs]
    if not new_docs:
        print("📚 새로 인덱싱할 문서 없음 (모든 문서가 이미 등록됨)")
        return

    for file in new_docs:
        print(f"➕ 새 문서 인덱싱: {file.name}")

        # 파일 내용 읽기
        if file.suffix == ".md":
            content = file.read_text(encoding="utf-8")
        elif file.suffix == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, data_only=True, read_only=True)  # 스타일 무시
                content_list = []

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        rows.append(row_str)
                    for i in range(0, len(rows), 5):
                        chunk_text = f"시트명: {sheet_name}\n" + "\n".join(rows[i:i+5])
                        content_list.append(chunk_text)
                content = "\n\n".join(content_list)
            except Exception as e:
                print(f"⚠️ {file.name} 읽기 실패: {e}")
                continue

        else:
            continue

        # 문단 단위 분리 및 청크 분할
        parts = [seg.strip() for seg in re.split(r'\n{2,}', content) if seg.strip()]
        for seg in parts:
            chunks = split_into_chunks(seg)
            for ci, chunk in enumerate(chunks, start=1):
                idx += 1
                ids.append(f"{file.name}__{idx}")
                texts.append(chunk)
                metadatas.append({"source": file.name, "chunk_idx": ci})

    if not texts:
        print("⚠️ 인덱싱할 문단 없음")
        return

    print("현재 저장된 문단 수:", chroma.count())

    # ✅ 임베딩 수행 및 배치 단위로 Chroma 추가
    embeddings = []
    for i in range(0, len(texts), BATCH_SIZE):
        batch_texts = texts[i:i + BATCH_SIZE]
        resp = openai_client.embeddings.create(model=EMBEDDING_MODEL, input=batch_texts)
        embeddings.extend([item.embedding for item in resp.data])

    MAX_CHROMA_BATCH = 5000
    for i in range(0, len(ids), MAX_CHROMA_BATCH):
        batch_ids = ids[i:i+MAX_CHROMA_BATCH]
        batch_texts = texts[i:i+MAX_CHROMA_BATCH]
        batch_embeds = embeddings[i:i+MAX_CHROMA_BATCH]
        batch_metas = metadatas[i:i+MAX_CHROMA_BATCH]

        chroma.add_documents(batch_ids, batch_texts, batch_embeds, batch_metas)
        print(f"🧩 Chroma에 {len(batch_ids)}개 문단 추가 완료 ({i + len(batch_ids)}/{len(ids)})")

    chroma.persist()
    print(f"✅ 새로 인덱싱된 문단 수: {len(ids)}개")


def query_and_answer(query: str, top_k=60):
    # 먼저 competition_handler로 수치 확인
    comp_res = competition_handler(query)
    if comp_res:
        if comp_res["type"] == "text":
            text_values = comp_res["content"]
            answer_lines = []
            for year, val in text_values.items():
                if val is None:
                    answer_lines.append(f"{year}년: 해당 연도의 경쟁률 정보를 찾을 수 없습니다.")
                else:
                    # ✅ 숫자든 문자열이든 안전하게 문자열로 변환
                    answer_lines.append(f"{year}년: {str(val)}")
            answer = "\n".join(answer_lines)
            # ✅ 최종 return 보장
            return {
                "type": "text",
                "answer": answer,
                "sources": []
            }
        else:
            return {
                "type": "graph",
                "answer": "그래프가 생성되었습니다.",
                "sources": []
            }

    # RAG 검색 (수치가 없는 경우만)
    q_emb_resp = openai_client.embeddings.create(
        model=EMBEDDING_MODEL,
        input=[query]
    )
    q_emb = q_emb_resp.data[0].embedding
    res = chroma.query(q_emb, top_k)
    docs = res.get("documents", [[]])[0]
    distances = res.get("distances", [[]])[0]

    RAG_THRESHOLD = 1.4
    if not docs or (len(distances) > 0 and distances[0] > RAG_THRESHOLD):
        # 문서와 무관 → 일반 LLM 반응
        general_prompt = f"다음 사용자 메시지에 자연스럽게 답변하세요.\n사용자 메시지: {query}\n"
        resp = openai_client.responses.create(
            model=LLM_MODEL,
            input=general_prompt,
            max_output_tokens=256
        )
        return getattr(resp, "output_text", str(resp))

    # 문서 기반 RAG 답변 (설명용)
    prompt = f"""
    너는 '안양대학교 공식 문서 기반 안내 챗봇'이다.
    아래 제공된 문서 내용만 근거로 사용해 답해라.

    작성 규칙(중요):
    - '1. 2. 3.' 같은 번호 나열 금지
    - 표를 그대로 복붙하지 말고, 사람이 말하듯 자연스럽게 요약
    - 먼저 1~2문장으로 전체 요약 → 그 다음 핵심만 3~6줄로 정리
    - 질문이 "뭐 있어?/뭐야?/알려줘" 같은 탐색형이면, 성격/카테고리로 묶어서 설명
    - 문서에 없는 내용은 추측하지 말고 "문서에 정보가 없습니다"라고 말해라
    - 답변은 한국어로, 과하게 길지 않게

    === 문서 내용 ===
    {chr(10).join(docs)}

    === 사용자 질문 ===
    {query}

    === 답변 ===
    """
    resp = openai_client.responses.create(
        model=LLM_MODEL,
        input=prompt,
        max_output_tokens=512
    )
    answer = getattr(resp, "output_text", str(resp))

    print("🔍 검색 결과 거리:", distances)
    if docs:
        print("🔍 가장 가까운 문서:", docs[0][:200], "...")
    print("🔍 쿼리 임베딩 dimension:", len(q_emb))

    return answer


ADMISSION_ALIAS = {
    "기회균형전형": "기회균형전형",
    "고른기회전형": "고른기회전형",
    "일반전형": "일반전형",
    "정시": "정시",
}


def competition_handler(query: str):
    parsed = parse_competition_query(query)

    print("🔹 parsed query:", parsed)

    # 필수값이 없으면 경쟁률 처리 대상 아님
    if not parsed.get("years") or not parsed.get("university") or not parsed.get("major"):
        return None

    results = {}

    # 정규화된 검색 문자열 만들기
    # major, university, admission 등을 검색용으로 정규화
    raw_univ = parsed.get("university")
    raw_major = parsed.get("major")
    raw_admission = parsed.get("admission")

    norm_univ = _normalize_for_search(raw_univ)
    # 학과명은 내부 공백이 섞여 들어오는 경우가 있으므로 모든 공백 제거
    norm_major = _normalize_for_search(raw_major)
    # 전형은 None일 수 있으니 안전하게 처리
    norm_admission = _normalize_for_search(raw_admission)

    # 전형 alias 적용 (alias 키도 정규화해서 매칭)
    # ADMISSION_ALIAS의 키가 한글 공백 포함 상태일 수 있으므로 normalize해서 확인
    alias_map = { _normalize_for_search(k): v for k, v in ADMISSION_ALIAS.items() }
    if norm_admission in alias_map:
        norm_admission = alias_map[norm_admission]

    for y in parsed.get("years", []):
        # find_competition_ratio에 전달할 때는 이미 정규화된 값을 사용
        try:
            val = find_competition_ratio(
                y,
                norm_univ,
                norm_major,
                norm_admission
            )
        except Exception as e:
            print(f"⚠️ find_competition_ratio 호출 중 예외: {e}")
            val = None

        # 숫자 포맷 통일
        if val is not None:
            try:
                val = round(float(val), 2)
            except Exception:
                # 숫자로 변환 불가하면 원래 값 그대로 두기
                pass

        results[y] = val

        print(f"🔹 year={y}, admission={norm_admission}, found val={val}")

    years = parsed.get("years", [])
    force_graph = len(years) >= 2

    # 그래프 요구 여부(또는 강제 그래프)
    if not parsed.get("wants_graph") and not force_graph:
        return {
            "type": "text",
            "content": results
        }

    img_b64 = generate_base64_graph(results)
    return {
        "type": "graph",
        "content": img_b64,
        "values": results
    }


# 직접 실행 시 문서 인덱싱
if __name__ == "__main__":
    print("문서 인덱싱 시작...")
    index_all_documents()
    print("문서 인덱싱 완료")
